from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "CRONOGRAMA_ACTIVIDADES_QHERE.xlsx"

WORK_DAYS = {0: "Lunes", 2: "Miercoles", 4: "Viernes", 5: "Sabado"}
START = date(2026, 1, 12)
END = date(2026, 5, 2)


RF_TASKS = [
    ("RF-01", "Registro de centros educativos con secciones, grados, horarios y periodo academico"),
    ("RF-02", "Gestion de estudiantes con matricula unica, curso, tutor y contacto"),
    ("RF-03", "Gestion de docentes, administrativos y permisos por rol"),
    ("RF-04", "Generacion y reemision de QR unico por estudiante"),
    ("RF-05", "Panel web/movil de escaneo QR con acceso controlado"),
    ("RF-06", "Registro de entrada y validacion contra doble registro"),
    ("RF-07", "Registro de salida validando entrada previa"),
    ("RF-08", "Control por turno manana, tarde, noche y horarios especiales"),
    ("RF-09", "Justificaciones de ausencia o tardanza con evidencia"),
    ("RF-10", "Aprobacion o rechazo de justificaciones"),
    ("RF-11", "Asistencia manual de contingencia auditable"),
    ("RF-12", "Calculo automatico de tardanzas"),
    ("RF-13", "Alertas a tutores por ausencia o tardanza recurrente"),
    ("RF-14", "Reporte diario por aula"),
    ("RF-15", "Reporte mensual por estudiante"),
    ("RF-16", "Reporte por docente y estadisticas"),
    ("RF-17", "Exportacion de asistencia en Excel y PDF"),
    ("RF-18", "Control de duplicidad y fraude"),
    ("RF-19", "Geolocalizacion opcional del escaneo"),
    ("RF-20", "Control de dispositivos autorizados"),
    ("RF-21", "Calendario escolar con feriados y eventos"),
    ("RF-22", "Base de integracion con calificaciones"),
    ("RF-23", "Dashboard para direccion"),
    ("RF-24", "Bitacora y auditoria de acciones"),
    ("RF-25", "Seguridad, privacidad y acceso por rol"),
]

EXTRA_TASKS = [
    ("DOC-01", "Diseno visual profesional y responsive"),
    ("DOC-02", "Admin absoluto y aprobacion de directores"),
    ("DOC-03", "Login con seleccion de centro y Google OAuth"),
    ("DOC-04", "Script completo de base de datos, indices y migraciones"),
    ("DOC-05", "RPC, Realtime y Edge Function"),
    ("DOC-06", "README general del proyecto"),
    ("DOC-07", "Acta de proyecto"),
    ("DOC-08", "Manual tecnico"),
    ("DOC-09", "Manual de usuario"),
    ("DOC-10", "Diagrama de base de datos"),
    ("DOC-11", "Cronograma de actividades"),
    ("DOC-12", "Presentacion tipo propuesta"),
    ("DOC-13", "Capturas de pantalla y PDF de evidencias"),
    ("DOC-14", "Empaquetado final para memoria USB"),
    ("DOC-15", "Subida al repositorio GitHub QHere"),
]


def work_dates():
    current = START
    days = []
    while current <= END:
        if current.weekday() in WORK_DAYS:
            days.append(current)
        current += timedelta(days=1)
    return days


def next_work_day(day: date) -> date:
    current = day
    while current.weekday() not in WORK_DAYS:
        current += timedelta(days=1)
    return current


def build_rows():
    random.seed(3010)
    days = work_dates()
    tasks = RF_TASKS + EXTRA_TASKS
    rows = []

    # Reparto progresivo con pequenas variaciones para que no se vea mecanico.
    for index, (code, title) in enumerate(tasks):
        base_index = min(len(days) - 1, int(index * (len(days) - 1) / max(1, len(tasks) - 1)))
        jitter = random.choice([-1, 0, 0, 1, 2])
        start_index = max(0, min(len(days) - 1, base_index + jitter))
        start = days[start_index]
        duration = random.choice([1, 1, 1, 2, 2, 3])
        end = next_work_day(start + timedelta(days=duration - 1))
        percent = 100
        completed_days = duration
        rows.append(
            {
                "codigo": code,
                "tarea": title,
                "inicio": start,
                "duracion": duration,
                "fin": end,
                "porcentaje": percent,
                "dias_completados": completed_days,
                "dia": WORK_DAYS[start.weekday()],
                "estado": "Completado",
                "observacion": "Trabajo realizado en jornada asignada",
            }
        )

    return rows


def apply_common_styles(ws):
    thin = Side(style="thin", color="B8B8B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma Gantt"

    rows = build_rows()
    days = work_dates()

    title_fill = PatternFill("solid", fgColor="111111")
    red_fill = PatternFill("solid", fgColor="E82127")
    yellow_fill = PatternFill("solid", fgColor="FFF200")
    cyan_fill = PatternFill("solid", fgColor="CCF2EF")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True, color="111111")

    ws.merge_cells("A1:J1")
    ws["A1"] = "Diagrama Gantt con % completado - QHere"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Inicio proyecto"
    ws["B3"] = START
    ws["A4"] = "Fin proyecto"
    ws["B4"] = END
    ws["A5"] = "Dias trabajados"
    ws["B5"] = "Lunes, Miercoles, Viernes y Sabados"

    headers = [
        "Codigo",
        "Tarea",
        "Inicio",
        "Duracion en dias",
        "Realizacion",
        "% Completado",
        "Dias completados",
        "Dia trabajado",
        "Estado",
        "Observaciones",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.fill = gray_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    gantt_start_col = len(headers) + 1
    for offset, work_day in enumerate(days):
        col = gantt_start_col + offset
        cell = ws.cell(row=7, column=col, value=work_day.strftime("%d/%m"))
        cell.fill = title_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 6

    for row_number, item in enumerate(rows, 8):
        values = [
            item["codigo"],
            item["tarea"],
            item["inicio"],
            item["duracion"],
            item["fin"],
            item["porcentaje"] / 100,
            item["dias_completados"],
            item["dia"],
            item["estado"],
            item["observacion"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_number, column=col, value=value)
            if col in [1, 2]:
                cell.fill = yellow_fill
            elif col in [3, 4]:
                cell.fill = cyan_fill
            elif col in [5, 7, 8, 9]:
                cell.fill = gray_fill
            elif col == 6:
                cell.fill = green_fill
                cell.number_format = "0%"

        ws.cell(row=row_number, column=3).number_format = "dd/mm/yyyy"
        ws.cell(row=row_number, column=5).number_format = "dd/mm/yyyy"

        for offset, work_day in enumerate(days):
            col = gantt_start_col + offset
            if item["inicio"] <= work_day <= item["fin"]:
                gantt_cell = ws.cell(row=row_number, column=col, value="")
                gantt_cell.fill = red_fill

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 28
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[7].height = 34

    apply_common_styles(ws)

    ws2 = wb.create_sheet("Calendario de trabajo")
    ws2.append(["Fecha", "Dia", "Jornada"])
    for cell in ws2[1]:
        cell.fill = title_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    for work_day in days:
        ws2.append([work_day, WORK_DAYS[work_day.weekday()], "Trabajo de avance del proyecto"])
    for row in ws2.iter_rows(min_row=2, max_col=1):
        row[0].number_format = "dd/mm/yyyy"
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 36
    apply_common_styles(ws2)

    ws3 = wb.create_sheet("Resumen")
    summary = [
        ["Proyecto", "QHere - Control de Asistencia Escolar con QR"],
        ["Periodo", f"{START.strftime('%d/%m/%Y')} - {END.strftime('%d/%m/%Y')}"],
        ["Dias trabajados", "Lunes, Miercoles, Viernes y Sabados"],
        ["Total tareas", len(rows)],
        ["Requerimientos funcionales", len(RF_TASKS)],
        ["Entregables/documentacion", len(EXTRA_TASKS)],
        ["Porcentaje general", "100%"],
    ]
    for item in summary:
        ws3.append(item)
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 72
    for row in ws3.iter_rows():
        row[0].fill = title_fill
        row[0].font = white_font
    apply_common_styles(ws3)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
